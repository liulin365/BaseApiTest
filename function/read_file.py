# 该模块主要用于从Excel文件读取测试用例

from os import path
from openpyxl import load_workbook
import yaml
import json


file_path = path.abspath(__file__)
target_case = path.dirname(path.dirname(file_path)) + '/file/测试用例.xlsx'
target_setting = path.dirname(path.dirname(file_path)) + '/file/default_setting.yaml'


class control_file:
    def __init__(self,sheet_name = None):
        self.sheet_name = sheet_name
        self.wb = load_workbook(target_case) # 载入文件
        self.ws = None

    def check_input(self):
        list_name = []
        if self.sheet_name == 'all' or self.sheet_name is None:
            for name in self.wb.sheetnames:
                list_name.append(name)
            return list_name
        else:
            list_name.append(self.sheet_name)
            return list_name

    def read_excel(self):
        sheet_name = self.check_input()
        data = []
        for name in sheet_name:
            self.ws = self.wb[name]
            for i in self.ws.iter_rows(min_row = 2, values_only=True):
                data.append(i)
        self.wb.close()
        return data

    def update_excel(self, sheet_name, row, column, value=None, json_key=None, json_value=None):
        """
        :param sheet_name: 工作表名称
        :param row: 行号（从1开始）
        :param column: 列号（数字）或列字母（如'E'）
        :param value: 直接写入的完整值（与json_key互斥）
        :param json_key: 要修改的JSON字段名（如"projectBuildingName"）
        :param json_value: 修改后的值
        """
        wb = load_workbook(target_case)
        ws = wb[sheet_name]

        # 定位单元格
        if isinstance(column, str):
            cell = ws[f"{column}{row}"]
        else:
            cell = ws.cell(row=row, column=column)

        # 直接覆盖整个单元格
        if value is not None:
            cell.value = value

        # 修改单元格内JSON的某个键值对
        elif json_key is not None:
            # 读取当前单元格的JSON字符串
            current_json_str = cell.value

            # 解析为字典
            current_dict = json.loads(current_json_str)

            # 修改指定键的值
            current_dict[json_key] = json_value

            # 写回单元格（转回JSON字符串）
            cell.value = json.dumps(current_dict, ensure_ascii=False)

        wb.save(target_case)
        wb.close()

    def read_yaml(self):
        with open(target_setting,mode='r',encoding='UTF-8') as f:
            data = yaml.safe_load(f)
        return data

    def update_yaml(self,first,masg,second = None):
        with open(target_setting,mode='r',encoding='UTF-8') as f:
            data = yaml.safe_load(f)
        data[first][second] = masg

        with open(target_setting, 'w', encoding='utf-8') as f:
            yaml.dump(data, f, allow_unicode=True, sort_keys=False)


    def str_to_dict(self,data): # 从Excel中读取出来的json有字符串引号，属于字符串类型
        json_str = data.strip("'")
        dict_data = json.loads(json_str)
        return dict_data



con_file = control_file()

if __name__ == '__main__':
    con_file = control_file()
    con_file.str_to_dict()









